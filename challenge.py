#!/usr/bin/env python3

from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime

def combine():
    wb = load_workbook('courses.xlsx')
    wb.create_sheet('combine')
    ws_time = wb['time']
    ws_students = wb['students']
    ws_combine = wb['combine']
    count = 1
    y_dict = {}
    for row in ws_students.values:
        if isinstance(row[0], datetime):
            if int(row[0].year) not in y_dict:
                y_dict[int(row[0].year)] = {}
            y_dict[int(row[0].year)][row[1]] = list(row)

    for row in ws_time.values:
        if isinstance(row[0], datetime):
            y_dict[int(row[0].year)][row[1]].append(row[2])
    ws_combine.append((ws_students['A1'].value, ws_students['B1'].value, ws_students['C1'].value, ws_time['C1'].value))
    for i in y_dict.values():
        for row in i.values():
            ws_combine.append(row)

    p = 0
    for i in ws_combine.values:
        print(i)
        if p > 3:
            break
        p += 1
    wb.save('courses.xlsx')
    return y_dict, list(map(value, ws_combine[1]))

def value(cell):
    return cell.value

def split():
    data, columns = combine()
    
    year = list(data.keys())
    for i in year:
        wb = Workbook()
        ws = wb.active
        ws.title = '{}'.format(i)
        ws.append(columns)
        for d in data[i].values():
            ws.append(d)
        wb.save('{}.xlsx'.format(i))

if __name__ == '__main__':
    combine()
    split()
